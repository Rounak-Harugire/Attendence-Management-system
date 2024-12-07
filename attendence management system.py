import tkinter as tk
import cv2
import os
import numpy as np
from PIL import Image
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# Main window setup
window = tk.Tk()
window.title("Attendance Management System using Face Recognition")
window.geometry('1280x720')
window.configure(background='lightgrey')

# Global dictionary to map IDs to student names and enrollments
id_to_name = {}
student_counter = 0  # Global counter for student IDs

# Function to capture and save images
def take_img():
    global student_counter
    enrollment = txt.get()
    name = txt2.get()
    subject_name = txt3.get()  # Get the subject name from the input field
    
    if enrollment == '' or name == '' or subject_name == '':
        Notification.configure(text="Enrollment, Name, and Subject are required!", bg="red", fg="white")
    else:
        if not os.path.exists("TrainingImage"):
            os.makedirs("TrainingImage")
        
        # Initialize camera and face detector
        cam = cv2.VideoCapture(0)
        if not cam.isOpened():
            Notification.configure(text="Unable to access the camera", bg="red", fg="white")
            return
        
        detector = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
        if detector.empty():
            Notification.configure(text="Face detector not loaded properly", bg="red", fg="white")
            return

        sampleNum = 0
        while True:
            ret, img = cam.read()
            if not ret:
                Notification.configure(text="Failed to grab frame from camera", bg="red", fg="white")
                break
            
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            faces = detector.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=5, minSize=(30, 30))

            for (x, y, w, h) in faces:
                sampleNum += 1
                # Draw rectangle around the face
                cv2.rectangle(img, (x, y), (x + w, y + h), (255, 0, 0), 2)
                
                # Save the captured face image in the 'TrainingImage' folder
                face_img = gray[y:y + h, x:x + w]
                face_path = f"TrainingImage/{student_counter}_{name}_{enrollment}_{subject_name}_{sampleNum}.jpg"
                cv2.imwrite(face_path, face_img)
                
            # Display the image with rectangles around faces
            cv2.imshow('Capturing Images', img)
            
            # Stop after 30 images or if 'q' is pressed
            if cv2.waitKey(1) & 0xFF == ord('q') or sampleNum >= 30:
                break
        
        # Release the camera and close the window
        cam.release()
        cv2.destroyAllWindows()

        # Update the student ID and store name, enrollment, and subject name
        id_to_name[student_counter] = (name, enrollment, subject_name)
        student_counter += 1  # Increment ID counter for the next student

        Notification.configure(text=f"Images Captured and Saved for Enrollment: {enrollment} Name: {name}", bg="green", fg="white")

# Function to train face recognition model
def train_image():
    recognizer = cv2.face.LBPHFaceRecognizer_create()
    detector = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')

    faces, ids = [], []
    for image_path in os.listdir('TrainingImage'):
        if image_path.endswith('.jpg'):
            img_path = os.path.join('TrainingImage', image_path)
            img = Image.open(img_path).convert('L')  # Convert to grayscale
            img_array = np.array(img, 'uint8')
            # Extract ID from the image filename
            id_ = int(image_path.split('_')[0])  # The first part of the filename is the ID
            faces.append(img_array)
            ids.append(id_)

    recognizer.train(faces, np.array(ids))
    if not os.path.exists('trainer'):
        os.makedirs('trainer')
    recognizer.save('trainer/trainer.yml')
    Notification.configure(text="Training Complete", bg="green", fg="white")

# Function to recognize faces, mark attendance, and store it in an Excel file
def recognize_face():
    recognizer = cv2.face.LBPHFaceRecognizer_create()
    recognizer.read('trainer/trainer.yml')
    detector = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')

    cam = cv2.VideoCapture(0)
    if not cam.isOpened():
        Notification.configure(text="Unable to access the camera", bg="red", fg="white")
        return

    font = cv2.FONT_HERSHEY_SIMPLEX
    marked_attendees = set()  # Set to keep track of already marked enrollments

    while True:
        ret, img = cam.read()
        if not ret:
            Notification.configure(text="Failed to grab frame from camera", bg="red", fg="white")
            break
        
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        faces = detector.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=5, minSize=(30, 30))
        for (x, y, w, h) in faces:
            id_, confidence = recognizer.predict(gray[y:y + h, x:x + w])
            if confidence < 100:  # lower value means better recognition
                name, enrollment, subject_name = id_to_name.get(id_, ("Unknown", "N/A", "N/A"))
                if enrollment not in marked_attendees and enrollment != "N/A":  # Check if not already marked
                    mark_attendance(name, enrollment, subject_name)
                    marked_attendees.add(enrollment)  # Add to marked attendees
            else:
                name, enrollment, subject_name = 'Unknown', 'N/A', 'N/A'

            # Display the name, enrollment, and subject on the image
            cv2.putText(img, f"Name: {name}, Enrollment: {enrollment}, Subject: {subject_name}", 
                        (x, y - 10), font, 0.8, (255, 255, 255), 2)
            cv2.rectangle(img, (x, y), (x + w, y + h), (255, 0, 0), 2)

        # Display the image with attendance info
        cv2.imshow('Face Recognition - Mark Attendance', img)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cam.release()
    cv2.destroyAllWindows()

    Notification.configure(text="Attendance Taken!", bg="green", fg="white")

# Function to mark attendance and store it in an Excel file
def mark_attendance(name, enrollment, subject_name):
    date_time = datetime.now()
    date = date_time.strftime('%Y-%m-%d')
    time = date_time.strftime('%H:%M:%S')
    
    # Load the existing Excel file if it exists, else create a new one
    if os.path.exists("attendance.xlsx"):
        wb = load_workbook("attendance.xlsx")  # Use load_workbook() instead of Workbook.load()
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Enrollment", "Name", "Subject", "Date", "Time"])  # Add header row
    
    ws = wb.active  # Select the active worksheet
    ws.append([enrollment, name, subject_name, date, time])  # Add a new row with attendance info
    
    # Save the file
    wb.save("attendance.xlsx")
    
    print(f"Attendance for {name} ({enrollment}) in {subject_name} on {date} at {time} saved.")  # Debugging line

# Function to clear text fields
def clear_enrollment():
    txt.delete(0, tk.END)

def clear_name():
    txt2.delete(0, tk.END)

def clear_subject():
    txt3.delete(0, tk.END)

# Function to show registered students
def check_registered_students_func():
    registered_students = "\n".join(f"ID: {key}, Name: {value[0]}, Enrollment: {value[1]}, Subject: {value[2]}" for key, value in id_to_name.items())
    Notification.configure(text=f"Registered Students:\n{registered_students}", bg="yellow", fg="black")

# Placeholder functions for new buttons
def train_image_func():
    train_image()

def automatic_attendance_func():
    recognize_face()

def manually_fill_attendance_func():
    Notification.configure(text="Manually Fill Attendance clicked.", bg="yellow", fg="black")

# GUI Elements
header = tk.Label(window, text="Attendance Management System using Face Recognition", bg="black", fg="white", width=70, height=2, font=('times', 20, 'bold'))
header.place(x=0, y=20)

# Entry fields and labels
lbl = tk.Label(window, text="Enter Enrollment:", fg="black", bg="lightgrey", font=('times', 14, 'bold'))
lbl.place(x=150, y=120)

txt = tk.Entry(window, width=30, font=('times', 14))
txt.place(x=400, y=120)

clear_btn1 = tk.Button(window, text="Clear", command=clear_enrollment, fg="white", bg="red", font=('times', 12, 'bold'))
clear_btn1.place(x=750, y=120)

lbl2 = tk.Label(window, text="Enter Name:", fg="black", bg="lightgrey", font=('times', 14, 'bold'))
lbl2.place(x=150, y=160)

txt2 = tk.Entry(window, width=30, font=('times', 14))
txt2.place(x=400, y=160)

clear_btn2 = tk.Button(window, text="Clear", command=clear_name, fg="white", bg="red", font=('times', 12, 'bold'))
clear_btn2.place(x=750, y=160)

lbl3 = tk.Label(window, text="Enter Subject:", fg="black", bg="lightgrey", font=('times', 14, 'bold'))
lbl3.place(x=150, y=200)

txt3 = tk.Entry(window, width=30, font=('times', 14))
txt3.place(x=400, y=200)

clear_btn3 = tk.Button(window, text="Clear", command=clear_subject, fg="white", bg="red", font=('times', 12, 'bold'))
clear_btn3.place(x=750, y=200)

# Buttons
take_img_btn = tk.Button(window, text="Capture Images", command=take_img, fg="black", bg="yellow", font=('times', 14, 'bold'))
take_img_btn.place(x=150, y=250)

train_img_btn = tk.Button(window, text="Train Images", command=train_image_func, fg="black", bg="pink", font=('times', 14, 'bold'))
train_img_btn.place(x=400, y=250)

automatic_btn = tk.Button(window, text="Automatic Attendance", command=automatic_attendance_func, fg="black", bg="orange", font=('times', 14, 'bold'))
automatic_btn.place(x=650, y=250)

check_registered_btn = tk.Button(window, text="Check Registered Students", command=check_registered_students_func, fg="white", bg="blue", font=('times', 14, 'bold'))
check_registered_btn.place(x=150, y=300)

clear_all_btn = tk.Button(window, text="Clear All", command=manually_fill_attendance_func, fg="white", bg="blue", font=('times', 14, 'bold'))
clear_all_btn.place(x=400, y=300)

Notification = tk.Label(window, text="", bg="lightgrey", fg="black", font=('times', 14, 'bold'))
Notification.place(x=150, y=350)

# Start the GUI
window.mainloop()
